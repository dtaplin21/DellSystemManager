import logging
import os
import fitz  # PyMuPDF for PDF processing
from typing import List, Dict, Any
import openpyxl  # for Excel processing

logger = logging.getLogger(__name__)

class DocumentProcessor:
    def __init__(self, openai_service):
        self.openai_service = openai_service
    
    def analyze_documents(self, file_paths: List[str], question: str = None) -> Dict[str, Any]:
        """
        Analyze multiple documents and generate insights based on their content
        
        Args:
            file_paths: List of paths to document files to analyze
            question: Specific question to focus the analysis on
            
        Returns:
            Dictionary containing analysis results
        """
        try:
            combined_text = ""
            
            for file_path in file_paths:
                file_ext = os.path.splitext(file_path)[1].lower()
                if file_ext == '.pdf':
                    text = self._extract_text_from_pdf(file_path)
                elif file_ext in ['.xlsx', '.xls']:
                    text = self._extract_text_from_excel(file_path)
                else:
                    logger.warning(f"Unsupported file type: {file_ext}")
                    continue
                
                # Add document content to combined text with separator
                if text:
                    combined_text += f"\n\n--- DOCUMENT: {os.path.basename(file_path)} ---\n\n"
                    combined_text += text
            
            if not combined_text:
                return {"error": "No text could be extracted from the provided documents"}
            
            # Default question if none provided
            if not question:
                question = "Provide a comprehensive analysis of these documents, including key findings, technical parameters, and any noteworthy information."
            
            # Get analysis from OpenAI
            analysis = self.openai_service.analyze_document_content(combined_text, question)
            
            # Structure the results
            return {
                "question": question,
                "analysis": analysis,
                "documentCount": len(file_paths),
                "processedFiles": [os.path.basename(path) for path in file_paths]
            }
            
        except Exception as e:
            logger.error(f"Error in analyze_documents: {str(e)}")
            return {"error": str(e)}
    
    def extract_data(self, file_path: str, extraction_type: str = 'auto') -> Dict[str, Any]:
        """
        Extract structured data from a document based on the extraction type
        
        Args:
            file_path: Path to the document file
            extraction_type: Type of extraction to perform (auto, qc_data, materials, tests)
            
        Returns:
            Dictionary containing extracted structured data
        """
        try:
            file_ext = os.path.splitext(file_path)[1].lower()
            
            if file_ext == '.pdf':
                text = self._extract_text_from_pdf(file_path)
            elif file_ext in ['.xlsx', '.xls']:
                # For Excel files, we'll extract based on the extraction type
                if extraction_type == 'qc_data':
                    return self._extract_qc_data_from_excel(file_path)
                else:
                    text = self._extract_text_from_excel(file_path)
            else:
                return {"error": f"Unsupported file type: {file_ext}"}
            
            if not text:
                return {"error": "No text could be extracted from the document"}
            
            # Use OpenAI to extract structured data based on extraction_type
            extraction_prompt = self._get_extraction_prompt(extraction_type)
            structured_data = self.openai_service.extract_structured_data(text, extraction_prompt)
            
            return {
                "extractionType": extraction_type,
                "data": structured_data,
                "sourceFile": os.path.basename(file_path)
            }
            
        except Exception as e:
            logger.error(f"Error in extract_data: {str(e)}")
            return {"error": str(e)}
    
    def _extract_text_from_pdf(self, file_path: str) -> str:
        """Extract text content from a PDF file"""
        try:
            doc = fitz.open(file_path)
            text = ""
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                text += page.get_text()
            return text
        except Exception as e:
            logger.error(f"Error extracting text from PDF: {str(e)}")
            return ""
    
    def _extract_text_from_excel(self, file_path: str) -> str:
        """Extract text content from an Excel file"""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            text = ""
            
            for sheet in wb.worksheets:
                text += f"\n--- SHEET: {sheet.title} ---\n"
                
                for row in sheet.iter_rows():
                    row_text = ""
                    for cell in row:
                        if cell.value is not None:
                            row_text += f"{cell.value}\t"
                    if row_text:
                        text += row_text.strip() + "\n"
            
            return text
        except Exception as e:
            logger.error(f"Error extracting text from Excel: {str(e)}")
            return ""
    
    def _extract_qc_data_from_excel(self, file_path: str) -> Dict[str, Any]:
        """
        Extract QC data specifically from an Excel file
        Attempts to automatically identify QC data columns and values
        """
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            qc_data = []
            
            # Keywords that might indicate QC data
            qc_keywords = [
                'panel', 'test', 'result', 'pass', 'fail', 'temperature', 
                'pressure', 'speed', 'technician', 'date', 'location'
            ]
            
            for sheet in wb.worksheets:
                # Try to find header row by looking for rows with QC keywords
                header_row = None
                headers = []
                
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    matches = 0
                    for cell in row:
                        if cell and any(keyword.lower() in str(cell).lower() for keyword in qc_keywords):
                            matches += 1
                    
                    # If we found multiple matches, this is likely a header row
                    if matches >= 2:
                        header_row = row_idx
                        headers = [str(cell).strip() if cell else f"Column_{i}" for i, cell in enumerate(row)]
                        break
                
                if not header_row:
                    # If no header found, use first row as header
                    header_row = 1
                    headers = [str(cell).strip() if cell else f"Column_{i}" for i, cell in enumerate(next(sheet.iter_rows(values_only=True)))]
                
                # Extract data rows
                for row in list(sheet.iter_rows(min_row=header_row+1, values_only=True)):
                    if any(cell is not None for cell in row):  # Skip empty rows
                        data_row = {}
                        for i, cell in enumerate(row):
                            if i < len(headers):
                                data_row[headers[i]] = cell
                        qc_data.append(data_row)
            
            # Identify important columns with more specific mapping
            processed_data = []
            for row in qc_data:
                processed_row = {}
                
                # Try to map general columns to specific field names
                for header, value in row.items():
                    header_lower = header.lower()
                    
                    if 'panel' in header_lower or 'location' in header_lower:
                        processed_row['panelId'] = value
                    elif 'date' in header_lower or 'time' in header_lower:
                        processed_row['date'] = value
                    elif 'test' in header_lower and 'type' in header_lower:
                        processed_row['type'] = value
                    elif 'technician' in header_lower or 'operator' in header_lower:
                        processed_row['technician'] = value
                    elif 'result' in header_lower or 'status' in header_lower or 'pass' in header_lower:
                        processed_row['result'] = value
                    elif 'note' in header_lower or 'comment' in header_lower:
                        processed_row['notes'] = value
                    elif 'temp' in header_lower:
                        processed_row['temperature'] = value
                    elif 'press' in header_lower:
                        processed_row['pressure'] = value
                    elif 'speed' in header_lower:
                        processed_row['speed'] = value
                    else:
                        # Include other columns as-is
                        processed_row[header] = value
                
                if processed_row:
                    processed_data.append(processed_row)
            
            # AI-assist to clean up and standardize the data
            if processed_data:
                processed_data = self.openai_service.clean_qc_data(processed_data)
            
            return {
                "extractionType": "qc_data",
                "data": processed_data,
                "sourceFile": os.path.basename(file_path),
                "rowsCount": len(processed_data)
            }
            
        except Exception as e:
            logger.error(f"Error extracting QC data from Excel: {str(e)}")
            return {"error": str(e)}
    
    def _get_extraction_prompt(self, extraction_type: str) -> str:
        """Return the appropriate extraction prompt based on type"""
        prompts = {
            'qc_data': """
                Extract all quality control data from this document. Look for:
                - Panel IDs or locations
                - Test types and procedures
                - Test results (pass/fail/values)
                - Dates and times
                - Technician names
                - Temperature, pressure, and welding speed values
                - Any notes or comments
                Return the data as a structured JSON array.
            """,
            'materials': """
                Extract all information about materials from this document. Look for:
                - Material types and specifications
                - Thickness values
                - Manufacturer information
                - Batch or lot numbers
                - Delivery dates
                - Installation dates
                Return the data as a structured JSON array.
            """,
            'tests': """
                Extract all test data from this document. Look for:
                - Test types
                - Test methods
                - Test results
                - Testing equipment used
                - Testing standards referenced
                - Acceptance criteria
                Return the data as a structured JSON array.
            """,
            'auto': """
                Analyze this document and extract all relevant information in a structured format.
                Identify the document type and extract appropriate information such as:
                - Test data
                - Material specifications
                - Project information
                - Quality control results
                - Personnel information
                Return the data as a structured JSON object with appropriate fields.
            """
        }
        
        return prompts.get(extraction_type, prompts['auto'])